from appium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from appium.options.android import UiAutomator2Options
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import os
from datetime import datetime
import time
from selenium.common.exceptions import WebDriverException

#append results in Excel sheet
def append_to_excel(sheet, row_data):
    sheet.append(row_data)
    row = sheet.max_row
    for col, value in enumerate(row_data, start=1):
        cell = sheet.cell(row=row, column=col)
        if value == "Pass":
            cell.font = Font(color='00FF00')
        elif value == "Fail":
            cell.font = Font(color='FF0000')

    workbook.save("results.xlsx")

# load Excel workbook and sheet
excel_file = "results.xlsx"
if os.path.exists(excel_file):
    workbook = load_workbook(excel_file)
    sheet = workbook.active
else:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "result.xlsx"
    # Add headers to the Excel sheet
    sheet.append(["Action", "Status", "Details"])
    
append_to_excel(sheet, ["restart time" + datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

#click button
def click_button(driver, xpath):
    try:
        wait = WebDriverWait(driver, 100)
        wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
        al = driver.find_element(By.XPATH, xpath)
        al.click()
        append_to_excel(sheet, ["Click Button", "Pass", xpath])
    except WebDriverException as e:
        append_to_excel(sheet, ["Click Button", "Fail", f"Failed to click the button: {e}"])

#enter text
def enter_text(driver, xpath, text):
    try:
        wait = WebDriverWait(driver, 100)
        wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
        el = driver.find_element(By.XPATH, xpath)
        el.send_keys(text)
        append_to_excel(sheet, ["Enter Text", "Pass", f"Entered '{text}' into {xpath}"])
    except WebDriverException as e:
        append_to_excel(sheet, ["Enter Text", "Fail", f"Failed to enter text: {e}"])

# check text
def check_text(driver, xpath, expected_text):
    try:
        wait = WebDriverWait(driver, 100)
        wait.until(EC.presence_of_element_located((By.XPATH, xpath)))
        el = driver.find_element(By.XPATH, xpath)
        actual_text = el.text
        if actual_text == expected_text:
            append_to_excel(sheet, ["Check Text", "Pass", f"Text '{actual_text}' matches expected text"])
        else:
            append_to_excel(sheet, ["Check Text", "Fail", f"Text '{actual_text}' does not match expected text '{expected_text}'"])
    except WebDriverException as e:
        append_to_excel(sheet, ["Check Text", "Fail", f"Failed to check text: {e}"])

options = UiAutomator2Options()
options.platform_name = 'Android'
options.device_name = 'emulator-5554'
options.app = '/Users/Akshat S/Downloads/base.apk'
options.automation_name = 'UiAutomator2'

appium_server_url = 'http://localhost:4723/wd/hub'

try:
    driver = webdriver.Remote(appium_server_url, options=options)
    print("Appium session started ")

    wait = WebDriverWait(driver, 100)
    time.sleep(5)

    allow_xpath = "//android.widget.Button[@resource-id='com.android.packageinstaller:id/permission_allow_button']"
    click_button(driver, allow_xpath)
    click_button(driver, allow_xpath)
    click_button(driver, allow_xpath)

    # >>>>Crate Account Details

    # CreateAccount_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/signUpButton']"
    # click_button(driver, CreateAccount_xpath)

    # Firstname_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox'])[1]"
    # enter_text(driver, Firstname_xpath, "Akshat")

    # Lastname_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox'])[2]"
    # enter_text(driver, Lastname_xpath, "Sharma")

    # Emailaddress_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox'])[3]"
    # enter_text(driver, Emailaddress_xpath, "shraks514@gmail.com")

    # Password_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox'])[4]"
    # enter_text(driver, Password_xpath, "Qwerty@123")

    # confirmpassword_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox'])[5]"
    # enter_text(driver, confirmpassword_xpath, "Qwerty@123")

    # eula_xpath = "(//android.widget.CheckBox[@resource-id='com.ciaqbroan.overture.mobile:id/termsCheckBox'])"
    # click_button(driver, eula_xpath)

    # checkbox_xpath = "(//android.widget.CheckBox[@resource-id='com.ciaqbroan.overture.mobile:id/conditionCheckBox'])"
    # click_button(driver, checkbox_xpath)

    # Createaccount_xpath = "//android.widget.Button[@resource-id='com.ciaqbroan.overture.mobile:id/btnRegister']"
    # click_button(driver, Createaccount_xpath)

    ##>>>>have to enter email code manually

    #>>>>login details

    textshown_xpath = "//android.widget.TextView[@text='Sign In']"
    expected_text = "Sign In"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/customeTitle' and @text='Email Address']"
    expected_text = "Email Address"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox' and @text='Email Address']"
    expected_text = "Email Address"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/customeTitle' and @text='Password']"
    expected_text = "Password"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox' and @text='Password']"
    expected_text = "Password"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/forgotPasswordView']"
    expected_text = "Forgot Password?"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.Button[@resource-id='com.ciaqbroan.overture.mobile:id/btnLogin']"
    expected_text = "Login"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/signUpButton']"
    expected_text = "Click here"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.TextView[@text='Select Language']"
    expected_text = "Select Language"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/languageTypeSpinner']"
    expected_text = "English"
    check_text(driver, textshown_xpath, expected_text)

    email_xpath = "//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox' and @text='Email Address']"
    enter_text(driver, email_xpath, "aksssh990@gmail.com")

    password_xpath = "//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox' and @text='Password']"
    enter_text(driver, password_xpath, "Qwerty@1234")

    login_xpath = "//android.widget.Button[@resource-id='com.ciaqbroan.overture.mobile:id/btnLogin']"
    click_button(driver, login_xpath)

    # notutorial1_xpath = "//android.widget.Button[@resource-id='com.ciaqbroan.overture.mobile:id/prev']"
    # click_button(driver, notutorial1_xpath)

    # notutorialcross_xpath = "//android.widget.ImageButton[@resource-id='com.ciaqbroan.overture.mobile:id/btnClose']"
    # click_button(driver, notutorialcross_xpath)
    # time.sleep(10)

    # ##>>>>home details

    # add_xpath = "(//android.widget.LinearLayout[@resource-id='com.ciaqbroan.overture.mobile:id/fab']/android.widget.ImageView)"
    # click_button(driver, add_xpath)

    # HomeName_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox' and @text='Home Name'])"
    # enter_text(driver, HomeName_xpath, "A's Home")

    # Area_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox' and @text='ex:2350'])"
    # enter_text(driver, Area_xpath, "2000")

    # BedroomNo_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox'])[3]"
    # enter_text(driver, BedroomNo_xpath, "4")

    # BathroomNo_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox'])[4]"
    # enter_text(driver, BathroomNo_xpath, "5")

    # FloorNo_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox'])[5]"
    # enter_text(driver, FloorNo_xpath, "2")

    # Zipcode_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox' and @text='Zip code'])"
    # enter_text(driver, Zipcode_xpath, "64642")

    # AddHome_xpath = "//android.widget.Button[@resource-id='com.ciaqbroan.overture.mobile:id/btnAddHome']"
    # click_button(driver, AddHome_xpath)

    # pop1_xpath = "//android.widget.ImageView[@resource-id='com.ciaqbroan.overture.mobile:id/imageViewShowCaseClose']"
    # click_button(driver, pop1_xpath)

    # pop2_xpath = "//android.widget.ImageButton[@resource-id='com.ciaqbroan.overture.mobile:id/btnClose']"
    # click_button(driver, pop2_xpath)
    # time.sleep(10)


    # ##>>>>to add room details

    # add_xpath = "(//android.widget.LinearLayout[@resource-id='com.ciaqbroan.overture.mobile:id/fab']/android.widget.ImageView)"
    # click_button(driver, add_xpath)
    
    # RoomName_xpath = "//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox']"
    # enter_text(driver, RoomName_xpath, "Room 1")

    # AddRoomBt_xpath = "//android.widget.Button[@resource-id='com.ciaqbroan.overture.mobile:id/btnAddRoom']"
    # click_button(driver, AddRoomBt_xpath)

    # pop3_xpath = "//android.widget.ImageView[@resource-id='com.ciaqbroan.overture.mobile:id/imageViewShowCaseClose']"
    # click_button(driver, pop3_xpath)

    # pop4_xpath = "//android.widget.ImageButton[@resource-id='com.ciaqbroan.overture.mobile:id/btnClose']"
    # click_button(driver, pop4_xpath)
    # time.sleep(5)
    
    ##>>>>Add new home

    # menu_xpath = "//android.widget.FrameLayout[@resource-id='com.ciaqbroan.overture.mobile:id/humberger']"
    # click_button(driver, menu_xpath)

    # myhomes_xpath = "(//android.widget.RelativeLayout[@resource-id='com.ciaqbroan.overture.mobile:id/rowView'])[3]"
    # click_button(driver, myhomes_xpath)

    # pop5_xpath = "//android.widget.Button[@resource-id='com.ciaqbroan.overture.mobile:id/prev']"
    # click_button(driver, pop5_xpath)

    # pop6_xpath = "//android.widget.ImageView[@resource-id='com.ciaqbroan.overture.mobile:id/imageViewShowCaseClose']"
    # click_button(driver, pop6_xpath)

    # AddNewHome_xpath = "(//android.widget.Button[@resource-id='com.ciaqbroan.overture.mobile:id/btnAddNewHome'])"
    # click_button(driver, AddNewHome_xpath)

    # HomeName_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox'])[1]"
    # enter_text(driver, HomeName_xpath, "B's Home")

    # Area_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox' and @text='ex:2350'])"
    # enter_text(driver, Area_xpath, "3000")

    # BedroomNo_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox'])[3]"
    # enter_text(driver, BedroomNo_xpath, "2")

    # BathroomNo_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox'])[4]"
    # enter_text(driver, BathroomNo_xpath, "2")

    # FloorNo_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox'])[5]"
    # enter_text(driver, FloorNo_xpath, "1")

    # Zipcode_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox' and @text='Zip code'])"
    # enter_text(driver, Zipcode_xpath, "64642")

    # AddHome_xpath = "//android.widget.Button[@resource-id='com.ciaqbroan.overture.mobile:id/btnAddHome']"
    # click_button(driver, AddHome_xpath)
    # time.sleep(5)

    textshown_xpath = "//android.widget.TextView[@text='To get Started, please click on the (+) icon below to create a home.']"
    expected_text = "To get Started, please click on the (+) icon below to create a home."
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/homeInfoTextView']"
    expected_text = "See the Quick Start guide for more information."
    check_text(driver, textshown_xpath, expected_text)

    # menu_xpath = "//android.widget.FrameLayout[@resource-id='com.ciaqbroan.overture.mobile:id/humberger']/android.widget.ImageView"
    # click_button(driver, menu_xpath)

    # Home_xpath = "(//android.widget.RelativeLayout[@resource-id='com.ciaqbroan.overture.mobile:id/rowView'])[1]"
    # click_button(driver, Home_xpath)

    menu_xpath = "//android.widget.FrameLayout[@resource-id='com.ciaqbroan.overture.mobile:id/humberger']/android.widget.ImageView"
    click_button(driver, menu_xpath)

    textshown_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/tvSettingsItem' and @text='Home']"
    expected_text = "Home"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/tvSettingsItem' and @text='Profile Settings']"
    expected_text = "Profile Settings"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/tvSettingsItem' and @text='My Homes']"
    expected_text = "My Homes"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/tvSettingsItem' and @text='My Devices']"
    expected_text = "My Devices"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/tvSettingsItem' and @text='My Users']"
    expected_text = "My Users"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/tvSettingsItem' and @text='Historical Data']"
    expected_text = "Historical Data"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/tvSettingsItem' and @text='App Settings']"
    expected_text = "App Settings"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/tvSettingsItem' and @text='Notification']"
    expected_text = "Notification"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/tvSettingsItem' and @text='Support']"
    expected_text = "Support"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/tvSettingsItem' and @text='Legal']"
    expected_text = "Legal"
    check_text(driver, textshown_xpath, expected_text)

    Home_xpath = "(//android.widget.RelativeLayout[@resource-id='com.ciaqbroan.overture.mobile:id/rowView'])[1]"
    click_button(driver, Home_xpath)

    add_xpath = "(//android.widget.LinearLayout[@resource-id='com.ciaqbroan.overture.mobile:id/fab']/android.widget.ImageView)"
    click_button(driver, add_xpath)

    textshown_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/customeTitle' and @text='Home Name']"
    expected_text = "Home Name"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox' and @text='Home Name']"
    expected_text = "Home Name"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/customeTitle' and @text='Approx Square Footage']"
    expected_text = "Approx Square Footage"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox' and @text='ex:2350']"
    expected_text = "ex:2350"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/customeTitle' and @text='# of Bedrooms']"
    expected_text = "# of Bedrooms"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox'])[3]"
    expected_text = "6"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/customeTitle' and @text='# of Bathrooms']"
    expected_text = "# of Bathrooms"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox'])[4]"
    expected_text = "6"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/customeTitle' and @text='# of Stories']"
    expected_text = "# of Stories"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox'])[5]"
    expected_text = "6"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.TextView[@resource-id='com.ciaqbroan.overture.mobile:id/customeTitle' and @text='Zip code']"
    expected_text = "Zip code"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox' and @text='Zip code']"
    expected_text = "Zip code"
    check_text(driver, textshown_xpath, expected_text)

    textshown_xpath = "//android.widget.Button[@resource-id='com.ciaqbroan.overture.mobile:id/btnAddHome']"
    expected_text = "Add Home"
    check_text(driver, textshown_xpath, expected_text)

    HomeName_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox' and @text='Home Name'])"
    enter_text(driver, HomeName_xpath, "A's Home")

    Area_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox' and @text='ex:2350'])"
    enter_text(driver, Area_xpath, "2000")

    BedroomNo_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox'])[3]"
    enter_text(driver, BedroomNo_xpath, "4")

    BathroomNo_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox'])[4]"
    enter_text(driver, BathroomNo_xpath, "5")

    FloorNo_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox'])[5]"
    enter_text(driver, FloorNo_xpath, "2")

    Zipcode_xpath = "(//android.widget.EditText[@resource-id='com.ciaqbroan.overture.mobile:id/customEditBox' and @text='Zip code'])"
    enter_text(driver, Zipcode_xpath, "64642")

    AddHome_xpath = "//android.widget.Button[@resource-id='com.ciaqbroan.overture.mobile:id/btnAddHome']"
    click_button(driver, AddHome_xpath)

    add_xpath = "(//android.widget.LinearLayout[@resource-id='com.ciaqbroan.overture.mobile:id/fab']/android.widget.ImageView)"
    click_button(driver, add_xpath)

    # pop1_xpath = "//android.widget.ImageView[@resource-id='com.ciaqbroan.overture.mobile:id/imageViewShowCaseClose']"
    # click_button(driver, pop1_xpath)

    # pop2_xpath = "//android.widget.ImageButton[@resource-id='com.ciaqbroan.overture.mobile:id/btnClose']"
    # click_button(driver, pop2_xpath)
    # time.sleep(10)

    # menu_xpath = "//android.widget.FrameLayout[@resource-id='com.ciaqbroan.overture.mobile:id/humberger']/android.widget.ImageView"
    # click_button(driver, menu_xpath)

    # ProfileSet_xpath = "(//android.widget.RelativeLayout[@resource-id='com.ciaqbroan.overture.mobile:id/rowView'])[2]"
    # click_button(driver, ProfileSet_xpath)

    # menu_xpath = "//android.widget.FrameLayout[@resource-id='com.ciaqbroan.overture.mobile:id/humberger']/android.widget.ImageView"
    # click_button(driver, menu_xpath)

    # myhomes_xpath = "(//android.widget.RelativeLayout[@resource-id='com.ciaqbroan.overture.mobile:id/rowView'])[3]"
    # click_button(driver, myhomes_xpath)

    # menu_xpath = "//android.widget.FrameLayout[@resource-id='com.ciaqbroan.overture.mobile:id/humberger']/android.widget.ImageView"
    # click_button(driver, menu_xpath)

    # MyDevices_xpath = "(//android.widget.RelativeLayout[@resource-id='com.ciaqbroan.overture.mobile:id/rowView'])[4]"
    # click_button(driver, MyDevices_xpath)

    # menu_xpath = "//android.widget.FrameLayout[@resource-id='com.ciaqbroan.overture.mobile:id/humberger']/android.widget.ImageView"
    # click_button(driver, menu_xpath)

    # MyUser_xpath = "(//android.widget.RelativeLayout[@resource-id='com.ciaqbroan.overture.mobile:id/rowView'])[5]"
    # click_button(driver, MyUser_xpath)

    # menu_xpath = "//android.widget.FrameLayout[@resource-id='com.ciaqbroan.overture.mobile:id/humberger']/android.widget.ImageView"
    # click_button(driver, menu_xpath)

    # HistoricalData_xpath = "(//android.widget.RelativeLayout[@resource-id='com.ciaqbroan.overture.mobile:id/rowView'])[6]"
    # click_button(driver, HistoricalData_xpath)

    # menu_xpath = "//android.widget.FrameLayout[@resource-id='com.ciaqbroan.overture.mobile:id/humberger']/android.widget.ImageView"
    # click_button(driver, menu_xpath)

    # AppSettings_xpath = "(//android.widget.RelativeLayout[@resource-id='com.ciaqbroan.overture.mobile:id/rowView'])[7]"
    # click_button(driver, AppSettings_xpath)

    # menu_xpath = "//android.widget.FrameLayout[@resource-id='com.ciaqbroan.overture.mobile:id/humberger']/android.widget.ImageView"
    # click_button(driver, menu_xpath)

    # notification_xpath = "(//android.widget.RelativeLayout[@resource-id='com.ciaqbroan.overture.mobile:id/rowView'])[8]"
    # click_button(driver, notification_xpath)

    # menu_xpath = "//android.widget.FrameLayout[@resource-id='com.ciaqbroan.overture.mobile:id/humberger']/android.widget.ImageView"
    # click_button(driver, menu_xpath)

    # support_xpath = "(//android.widget.RelativeLayout[@resource-id='com.ciaqbroan.overture.mobile:id/rowView'])[9]"
    # click_button(driver, support_xpath)

    # menu_xpath = "//android.widget.FrameLayout[@resource-id='com.ciaqbroan.overture.mobile:id/humberger']/android.widget.ImageView"
    # click_button(driver, menu_xpath)

    # legal_xpath = "(//android.widget.RelativeLayout[@resource-id='com.ciaqbroan.overture.mobile:id/rowView'])[10]"
    # click_button(driver, support_xpath)

    # menu_xpath = "//android.widget.FrameLayout[@resource-id='com.ciaqbroan.overture.mobile:id/humberger']/android.widget.ImageView"
    # click_button(driver, menu_xpath)

    ##>>>>delete account

    menu_xpath = "//android.widget.FrameLayout[@resource-id='com.ciaqbroan.overture.mobile:id/humberger']"
    click_button(driver, menu_xpath)

    ProfileSet_xpath = "(//android.widget.RelativeLayout[@resource-id='com.ciaqbroan.overture.mobile:id/rowView'])[2]"
    click_button(driver, ProfileSet_xpath)

    logout_xpath = "//android.widget.RelativeLayout[@resource-id='com.ciaqbroan.overture.mobile:id/logoutViewLayout']"
    click_button(driver, logout_xpath)

    # DeleteAccBT_xpath = "//android.widget.RelativeLayout[@resource-id='com.ciaqbroan.overture.mobile:id/removeAcountViewLayout']"
    # click_button(driver, DeleteAccBT_xpath)

    # FinalDelete_xpath = "//android.widget.Button[@resource-id='com.ciaqbroan.overture.mobile:id/buttonPositive']"
    # click_button(driver, FinalDelete_xpath)

    # deletedone_xpath = "//android.widget.Button[@resource-id='com.ciaqbroan.overture.mobile:id/buttonPositive']"
    # click_button(driver, deletedone_xpath)
    time.sleep(15)

    driver.back()

except Exception as e:
    print(f"Error starting Appium session: {e}")

"""finally:
    if 'driver' in locals():
        driver.quit()
        print("Appium session closed.")
    else:
        print("No Appium session to close.")"""
